# 필요한 모듈 임포트
from flask import Flask, render_template, Response, send_file  # Flask 웹 프레임워크 관련 모듈
import cv2  # OpenCV 이미지 처리 라이브러리
import numpy as np  # 다차원 배열을 다루는 NumPy 라이브러리
from os import listdir  # 디렉터리 내 파일 목록을 얻기 위한 os 모듈 함수
from os.path import isfile, join  # 파일 경로 조합 및 파일 존재 여부 확인을 위한 os.path 모듈 함수
from scipy.spatial import distance  # 공간 거리 계산을 위한 SciPy 라이브러리
from imutils import face_utils  # 얼굴 유틸리티 함수를 제공하는 Imutils 라이브러리
import dlib  # 얼굴 감지 및 특징점 예측을 위한 dlib 라이브러리
from pygame import mixer  # 음악 재생을 위한 Pygame 라이브러리
import pyttsx3  # 텍스트 음성 변환을 위한 pyttsx3 라이브러리
import time  # 시간 관련 작업을 위한 내장 time 모듈
from ultralytics import YOLO  # YOLO 객체 검출을 위한 Ultralytics YOLO 라이브러리
import pygame.mixer  # 음악 재생을 위한 Pygame 라이브러리
import os  # 운영 체제와 상호 작용하기 위한 내장 os 모듈
from openpyxl import Workbook, load_workbook  # 엑셀 파일 작성 및 로드를 위한 OpenPyXL 라이브러리
import datetime  # 날짜 및 시간 관련 작업을 위한 내장 datetime 모듈
import openpyxl  # 엑셀 파일 읽기를 위한 OpenPyXL 라이브러리
from io import BytesIO  # 메모리 내의 이진 데이터를 처리하기 위한 IO 모듈의 BytesIO 클래스
import pandas as pd  # 데이터 프레임 생성 및 조작을 위한 Pandas 라이브러리

# 필요한 환경 변수 설정
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'

# Flask 애플리케이션 초기화
app = Flask(__name__)

# 웹캠 영상 캡처를 위한 VideoCapture 객체 초기화
cap = cv2.VideoCapture(0)



#######################################################################
#################LBPH 얼굴 인식기를 이용한 얼굴 감지 시스템##############
######################################################################


# 이미지 데이터가 저장된 경로
data_path = 'static/faces/'

# data_path에 있는 파일 중 파일인 것들만 리스트로 가져옴
onlyfiles = [f for f in listdir(data_path) if isfile(join(data_path, f))]

# 얼굴 인식을 위한 훈련 데이터와 레이블 초기화
Training_Data, Labels = [], []

# 모든 파일에 대해 반복
for i, files in enumerate(onlyfiles):
    # 이미지 파일 경로
    image_path = data_path + onlyfiles[i]
    
    # 이미지를 흑백 이미지로 읽어들임
    images = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    
    # 이미지 데이터를 uint8 형식으로 변환하여 Training_Data에 추가
    Training_Data.append(np.asarray(images, dtype=np.uint8))
    
    # 레이블에 인덱스 i를 추가
    Labels.append(i)

# Labels를 numpy 배열로 변환
Labels = np.asarray(Labels, dtype=np.int32)

# LBPH 얼굴 인식기 생성
model = cv2.face.LBPHFaceRecognizer_create()

# 훈련 데이터로 모델을 학습
model.train(np.asarray(Training_Data), np.asarray(Labels))

# 모델 학습 완료 메시지 출력
print("모델 학습 끝 ~~~~~")

# 얼굴을 인식하기 위한 하르 분류기 불러오기
face_classifier = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

# 음악 재생을 위한 mixer 초기화
mixer.init()

# 재생할 음악 파일 로드
mixer.music.load(r"D:\bigdata_itwill\final_project\data\sound\music.wav")

# 이미지에서 얼굴을 감지하는 함수 정의
def face_detector(img, size=0.5):
    # 이미지를 흑백으로 변환
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # 얼굴을 감지하고 좌표 반환
    faces = face_classifier.detectMultiScale(gray, 1.3, 5)

    # 얼굴이 없을 경우 처리
    if faces is ():
        return img, []

    # 얼굴이 있는 경우에는 각 얼굴 주변에 사각형 그리기
    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 255), 2)
        # 얼굴 ROI(Region of Interest) 추출 및 크기 조정
        roi = img[y:y + h, x:x + w]
        roi = cv2.resize(roi, (200, 200))

    return img, roi


########################################################################
####################실시간 얼굴 감지와 데이터 저장 프로세스################
########################################################################

def save_to_excel(data, file_name):
    try:
        # 기존 워크북을 불러옵니다.
        wb = load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        # 파일이 없으면 새로운 워크북을 생성합니다.
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Date', 'Time', 'Name'])  # 새 파일에 헤더를 추가합니다.

    # 데이터를 시트에 추가합니다.
    sheet.append(data)
    wb.save(file_name)

# 현재 날짜와 시간을 문자열로 변환합니다.
current_time = datetime.datetime.now()
date_str = current_time.strftime("%Y-%m-%d")
time_str = current_time.strftime("%H:%M:%S")

# 새로운 엑셀 파일명을 생성합니다.
excel_file_name = f'event_data.xlsx'

# 데이터 준비
# 이벤트 상태 초기화
event_status = 'UNKNOWN'
event_status1 = 'USER'


# 적절한 시간 간격을 설정합니다. 예: 5초마다 저장
save_interval = 4  # 초 단위

def generate_frames():
    # 마지막으로 저장된 시간을 추적하기 위한 변수를 초기화합니다.
    last_save_time = time.time()  # 루프 내에서 초기화

    while True:
        # 비디오 프레임을 읽어옵니다.
        success, frame = cap.read()
        
        # 비디오 프레임을 성공적으로 읽었는지 확인합니다.
        if not success:
            break
        else:
            # 얼굴을 감지하고 프레임에 얼굴을 표시합니다.
            img, face = face_detector(frame)
            
            try:
                # 얼굴을 그레이스케일로 변환합니다.
                face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
                
                # 얼굴 인식 모델을 사용하여 얼굴을 예측합니다.
                result = model.predict(face)

                # 신뢰도를 계산하고 화면에 표시합니다.
                if result[1] < 500:
                    confidence = int(100 * (1 - (result[1]) / 300))
                    display_string = str(confidence) + '%'
                cv2.putText(img, display_string, (100, 120), cv2.FONT_HERSHEY_COMPLEX, 1, (250, 120, 255), 2)

                # 현재 시간을 가져옵니다.
                current_time = datetime.datetime.now()

                # 신뢰도가 일정 수준 이상인 경우 상태를 업데이트하고 데이터를 저장합니다.
                if confidence > 78:
                    cv2.putText(img, "Unlocked", (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
                    data_to_save = [current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S"), event_status1]
                else:
                    cv2.putText(img, "Locked", (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
                    data_to_save = [current_time.strftime("%Y-%m-%d"), current_time.strftime("%H:%M:%S"), event_status]

                # 일정 시간 간격으로만 데이터를 저장합니다.
                if current_time.timestamp() - last_save_time >= save_interval:
                    save_to_excel(data_to_save, excel_file_name)
                    last_save_time = current_time.timestamp()

            except:
                # 얼굴이 감지되지 않은 경우 예외를 처리합니다.
                cv2.putText(img, "Face Not Found", (250, 450), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 0), 2)
                pass

            # 프레임을 JPEG 형식으로 인코딩하고 전송합니다.
            ret, buffer = cv2.imencode('.jpg', img)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')


#############################################################
#########눈 감김 감지 및 경고 시스템###########################
#############################################################

def eye_aspect_ratio(eye):
    # 눈의 세 가지 길이를 계산합니다.
    A = distance.euclidean(eye[1], eye[5])  # 수직 길이
    B = distance.euclidean(eye[2], eye[4])  # 수직 길이
    C = distance.euclidean(eye[0], eye[3])  # 수평 길이
    
    # 눈의 종횡비(AspectRatio)를 계산합니다.
    ear = (A + B) / (2.0 * C)
    return ear

# 눈 감지 임계값을 설정합니다.
thresh = 0.13

# 얼굴 검출기와 눈 예측기를 초기화합니다.
detect = dlib.get_frontal_face_detector()
predict = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")

# 왼쪽 눈과 오른쪽 눈을 지정합니다.
(lStart, lEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["left_eye"]
(rStart, rEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["right_eye"]

# 눈 감지를 포함하는 프레임 생성 함수입니다.
def generate_frames_with_eye_detection():
    flag = 0
    # 음성 출력 엔진을 초기화합니다.
    eye_engine = pyttsx3.init()
    # 사운드 재생을 위한 파이게임 라이브러리를 초기화합니다.
    pygame.mixer.init()
    
    while True:
        # 비디오 프레임을 읽어옵니다.
        success, frame = cap.read()
        if not success:
            break
        else:
            # 프레임을 흑백으로 변환합니다.
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            
            # 얼굴을 검출합니다.
            subjects = detect(gray, 0)
            
            # 각 얼굴에 대해 반복합니다.
            for subject in subjects:
                # 얼굴 형상을 예측합니다.
                shape = predict(gray, subject)
                shape = face_utils.shape_to_np(shape)
                leftEye = shape[lStart:lEnd]  # 왼쪽 눈 형상을 가져옵니다.
                rightEye = shape[rStart:rEnd]  # 오른쪽 눈 형상을 가져옵니다.
                leftEAR = eye_aspect_ratio(leftEye)  # 왼쪽 눈 종횡비를 계산합니다.
                rightEAR = eye_aspect_ratio(rightEye)  # 오른쪽 눈 종횡비를 계산합니다.
                ear = (leftEAR + rightEAR) / 2.0  # 양쪽 눈의 평균 종횡비를 계산합니다.
                leftEyeHull = cv2.convexHull(leftEye)
                rightEyeHull = cv2.convexHull(rightEye)
                
                # 눈 주변에 윤곽선을 그립니다.
                cv2.drawContours(frame, [leftEyeHull], -1, (0, 255, 0), 1)
                cv2.drawContours(frame, [rightEyeHull], -1, (0, 255, 0), 1)
                
                # 눈 감지 임계값을 초과하는 경우 알림을 표시하고 소리를 재생합니다.
                if ear < thresh:
                    flag += 1  # 졸음 카운터를 증가시킵니다.
                    cv2.putText(frame, "WAKE UP", (10, 30),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)  # 화면에 "WAKE UP" 텍스트 표시
                    mixer.music.play()  # 소리를 재생합니다.
                    
                    if flag >= 4:  # 졸음이 연속으로 감지된 경우
                        print("졸음 감지 소리가 3번 이상 울렸습니다. 졸음 감지 상태입니다.")
                        cv2.putText(frame, "wake up bitches!!", (10, 30),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                        eye_engine.say("주의 졸음이 감지되었습니다.")
                        eye_engine.say("가까운 휴게소나 졸음쉬터에서 차를 세우고 휴식을 취하십시오")
                        eye_engine.runAndWait()
                        
                        # 졸음이 감지되면 소리를 재생합니다.
                        pygame.mixer.music.load(r'D:\bigdata_itwill\final_project\data\sound\music.wav')  # 소리 파일의 경로로 변경
                        pygame.mixer.music.play()
                else:
                    flag = 0  # 졸음 카운터를 초기화합니다.
                    
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')


#######################################################################
#####YOLOv8를 활용한 졸음 탐지 및 경고 시스템#############################
########################################################################

def generate_yolo_frames():
    # YOLOv8 모델을 불러옵니다
    yolo_model = YOLO(r"D:\bigdata_itwill\final_project\yolov8\yolov8\runs\detect\train\weights\best.pt")
    
    # pyttsx3 엔진을 초기화합니다
    yolo_engine = pyttsx3.init()
    
    # Pygame mixer 초기화
    pygame.mixer.init()
    
    # 졸음 감지 횟수를 저장하기 위한 변수를 초기화합니다
    counter = 0
    full_count = 0
    
    while True:
        # 비디오 프레임을 읽어옵니다.
        success, frame = cap.read()

        if success:
            # YOLOv8를 사용하여 프레임에서 물체를 탐지합니다.
            yolo_results = yolo_model.track(frame, persist=True)

            for item in yolo_results:
                # 각 객체에 대해 클래스와 신뢰도를 가져와서 확인합니다.
                for cls, conf in zip(item.boxes.cls, item.boxes.conf):
                    if cls == 16.0 and conf > 0.50:
                        counter += 1  # 졸음 탐지 카운터를 증가시킵니다.
                        print('카운트', counter)
                    elif cls == 15.0:
                        counter = 0  # 카운터를 초기화합니다.
                        print('카운트', counter)
                    
                    # 일정 횟수 이상 졸음이 감지되면 알람을 울립니다.
                    if counter > 6:
                        print('알람')
                        full_count += 1  # 전체 졸음 감지 횟수를 증가시킵니다.
                        print('풀카운트', full_count)
                        
                        # 음성으로 졸음 감지 메시지를 출력합니다.
                        yolo_engine.say("주의: 졸음이 감지되었습니다.")
                        yolo_engine.runAndWait()

                        # 졸음이 감지되면 소리를 재생합니다.
                        pygame.mixer.music.load(r'D:\bigdata_itwill\final_project\data\sound\alarm.wav')  # 소리 파일의 경로로 변경
                        pygame.mixer.music.play()
                        
                        counter = 0  # 졸음 감지 카운터를 초기화합니다.
                        
                        # 일정 횟수 이상 졸음이 감지되면 안전 메시지를 출력합니다.
                        if full_count > 2:
                            time.sleep(5)  # 일정 시간 동안 대기합니다.
                            yolo_engine.say("가까운 휴게소나 졸음쉬터에서 차를 세우고 휴식을 취하십시오")
                            yolo_engine.runAndWait()

                            # 휴식을 취하라는 메시지가 출력되면 졸음 감지 소리를 재생합니다.
                            pygame.mixer.music.load(r'D:\bigdata_itwill\final_project\data\sound\sound1.wav')  # 소리 파일의 경로로 변경
                            pygame.mixer.music.play()

            print('='*40)

            # 결과를 프레임에 시각화하여 표시합니다.
            annotated_frame = yolo_results[0].plot()

            # 어노테이션된 프레임을 표시합니다.
            cv2.imshow("YOLOv8 추적", annotated_frame)

            ret, buffer = cv2.imencode('.jpg', annotated_frame)
            frame = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')
        else:
            break




############################################################################
############Flask 애플리케이션의 라우팅 및 기능 정의##########################
###########################################################################

# Flask 애플리케이션의 루트 경로를 설정하고 main.html을 렌더링합니다.
@app.route('/')
def main():
    return render_template('main.html')

# 첫 번째 페이지의 경로를 설정하고 first.html을 렌더링합니다.
@app.route('/first_page')
def first_page():
    return render_template('first.html')

# 두 번째 페이지의 경로를 설정하고 drowsy12.html을 렌더링합니다.
@app.route('/one_page')
def one_page():
    return render_template('drowsy12.html')


# landmark 페이지 입니다.
@app.route('/second_page')
def second_page():
    return render_template('second.html')


# 랜드마크 페이지의 경로를 설정하고 landmark.html을 렌더링합니다.
@app.route('/landmark')
def landmark():
    return render_template('landmark.html')

# YOLO 페이지의 경로를 설정하고 yolo.html을 렌더링합니다.
@app.route('/yolo')
def yolo():
    return render_template('yolo.html')

# 비디오 피드의 경로를 설정하고 generate_frames() 함수로부터 비디오 프레임을 반환합니다.
@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

# 눈 감지를 포함한 비디오 피드의 경로를 설정하고 generate_frames_with_eye_detection() 함수로부터 비디오 프레임을 반환합니다.
@app.route('/video_feed_with_eye_detection')
def video_feed_with_eye_detection():
    return Response(generate_frames_with_eye_detection(), mimetype='multipart/x-mixed-replace; boundary=frame')

# YOLO 비디오 피드의 경로를 설정하고 generate_yolo_frames() 함수로부터 비디오 프레임을 반환합니다.
@app.route('/yolo_video_feed')
def yolo_video_feed():
    return Response(generate_yolo_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


# 엑셀 다운로드의 경로를 설정하고 엑셀 파일을 다운로드합니다.
@app.route('/download_excel')
def download_excel():
    # 엑셀 파일을 데이터프레임으로 읽기
    df = pd.read_excel(excel_file_name)

    # BytesIO를 사용하여 메모리에 엑셀 파일을 생성
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)

    # 엑셀 파일 다운로드
    return send_file(excel_buffer, download_name=excel_file_name, as_attachment=True)

# Flask 애플리케이션을 실행합니다.
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
    
    
    
    
    
    
    
    
    
    
    
    
